from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from urllib.parse import quote_plus
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from dotenv import load_dotenv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# --- CONFIG ---
JOB_KEYWORD = "Mechatronics Engineer"
COUNTRIES = ["Turkey", "Türkiye"]
DATE_POSTED = "24h"
EXPERIENCE_LEVELS = []
WORKPLACE_TYPES = []

# DATE_POSTED codes:            # EXPERIENCE_LEVELS codes:          # WORKPLACE_TYPES codes:    
# "any" = Any time              # [] = All levels                   # [] = All types       
# "24h" = Past 24 hours         # ["1"] = Internship                # ["1"] = On-site  
# "week" = Past week            # ["2"] = Entry level               # ["2"] = Remote   
# "month" = Past month          # ["3"] = Associate                 # ["3"] = Hybrid   
                                # ["4"] = Mid-Senior level          
                                # ["5"] = Director                  
                                # ["6"] = Executive                 
           
# --- EMAIL (.env dosyasından okunur) ---
SENDER_EMAIL = os.getenv("SENDER_EMAIL")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")  # Google App Password
RECEIVER_EMAIL = os.getenv("RECEIVER_EMAIL")

# --- SCROLL ---
MAX_SCROLL_ATTEMPTS = 200
SCROLL_PAUSE = 5
DETAIL_PAUSE = 2

# --- SAFE FILENAMES ---
safe_keyword   = JOB_KEYWORD.replace(" ", "_")
safe_exp       = "_".join(EXPERIENCE_LEVELS) if EXPERIENCE_LEVELS else "all"
safe_workplace = "_".join(WORKPLACE_TYPES)   if WORKPLACE_TYPES   else "all"
safe_date      = DATE_POSTED

# --- SETUP CHROME ---
options = Options()
options.add_argument("--start-maximized")
options.add_argument("--incognito")
options.add_argument("--log-level=3")
options.add_experimental_option("excludeSwitches", ["enable-logging"])
driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options
)

# --- HELPER FUNCTIONS ---
def build_linkedin_url(keyword, location, exp_levels, workplace_types, date_posted):
    exp_param       = ",".join(exp_levels)       if exp_levels       else ""
    workplace_param = ",".join(workplace_types)  if workplace_types  else ""
    date_param = ""
    if date_posted == "24h":     date_param = "r86400"
    elif date_posted == "week":  date_param = "r604800"
    elif date_posted == "month": date_param = "r2592000"

    url = f"https://www.linkedin.com/jobs/search/?keywords={quote_plus(keyword)}&location={quote_plus(location)}"
    if exp_param:       url += f"&f_E={exp_param}"
    if workplace_param: url += f"&f_WT={workplace_param}"
    if date_param:      url += f"&f_TPR={date_param}"
    url += "&position=1&pageNum=0"
    return url

def scroll_page(driver):
    attempt     = 0
    last_height = driver.execute_script("return document.body.scrollHeight")
    while attempt < MAX_SCROLL_ATTEMPTS:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE)
        try:
            btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CLASS_NAME, "infinite-scroller__show-more-button"))
            )
            btn.click()
            time.sleep(SCROLL_PAUSE)
        except:
            pass
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height
        attempt += 1

def fetch_job_details(job_url):
    job_desc = company_desc = ""
    if not job_url:
        return job_desc, company_desc
    try:
        driver.get(job_url)
        time.sleep(DETAIL_PAUSE)
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CLASS_NAME, "description__text"))
        )
        soup        = BeautifulSoup(driver.page_source, "html.parser")
        job_div     = soup.find("div", class_="description__text")
        job_desc    = job_div.get_text(separator="\n", strip=True) if job_div else ""
        company_div = soup.find("div", class_="show-more-less-html__markup")
        company_desc= company_div.get_text(separator="\n", strip=True) if company_div else ""
    except Exception as e:
        print(f"⚠️ İş detayı alınamadı: {e}")
    return job_desc, company_desc

def save_to_excel(jobs, filename):
    header_fill  = PatternFill("solid", start_color="1B3A6B", end_color="1B3A6B")
    alt_row_fill = PatternFill("solid", start_color="EEF2F7", end_color="EEF2F7")
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell_font    = Font(name="Arial", size=10)
    link_font    = Font(name="Arial", size=10, color="0563C1", underline="single")
    thin         = Side(style="thin", color="D0D7E3")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center       = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left         = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    columns = [
        ("Ülke",               12),
        ("İş Başlığı",         35),
        ("Şirket Adı",         28),
        ("Şirket URL",         40),
        ("Konum",              25),
        ("Yan Haklar",         20),
        ("İlan Tarihi",        15),
        ("Tarama Tarihi",      17),   # hangi gün eklendiği
        ("İş İlanı URL",       45),
        ("İş Açıklaması",      60),
        ("Şirket Açıklaması",  60),
    ]

    # --- Dosya varsa aç, yoksa yeni oluştur ---
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        start_row = ws.max_row + 1   # son satırın bir altından başla
        is_new    = False
        print(f"📂 Mevcut dosya açıldı: {filename} (son satır: {ws.max_row})")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title  = "LinkedIn İlanları"
        start_row = 2
        is_new    = True

        # Başlık satırını sadece yeni dosyada yaz
        for col_idx, (header, width) in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes    = "A2"

    # --- Mükerrer kontrol: mevcut URL'leri topla ---
    existing_urls = set()
    if not is_new:
        for row in ws.iter_rows(min_row=2, values_only=True):
            url_val = row[8] if len(row) > 8 else None   # 9. sütun = İş İlanı URL
            if url_val:
                existing_urls.add(url_val.strip())

    scrape_date = datetime.now().strftime("%d.%m.%Y %H:%M")
    added       = 0
    skipped     = 0

    for job in jobs:
        job_url = job.get("job_url", "").strip()

        # Daha önce eklenmişse geç
        if job_url and job_url in existing_urls:
            skipped += 1
            continue

        row_idx = start_row + added
        fill    = alt_row_fill if row_idx % 2 == 0 else None

        values = [
            job.get("country", ""),
            job.get("job_title", ""),
            job.get("company_name", ""),
            job.get("company_url", ""),
            job.get("location", ""),
            job.get("benefit", ""),
            job.get("posted", ""),
            scrape_date,                         # Tarama Tarihi
            job_url,
            job.get("job_description", ""),
            job.get("company_description", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = border
            cell.alignment = left
            cell.font      = link_font if col_idx in (4, 9) and value else cell_font
            if fill:
                cell.fill  = fill
        ws.row_dimensions[row_idx].height = 80

        existing_urls.add(job_url)
        added += 1

    # Filtre aralığını güncelle
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"

    wb.save(filename)
    print(f"📊 {added} yeni ilan eklendi | {skipped} mükerrer atlandı → {filename}")

def send_job_email(jobs, sender, receiver, password):
    if not jobs:
        print("⚠️ Gönderilecek ilan yok.")
        return False
    jobs_html = ""
    for job in jobs:
        jobs_html += f"""
        <div style="margin:10px 0;padding:12px;border-left:4px solid #1B3A6B;background:#f7f9fc;border-radius:5px;">
            <strong style="font-size:15px;">{job['job_title']}</strong><br>
            <em>{job['company_name']}</em> — {job['location']}<br>
            <a href="{job['job_url']}" target="_blank" style="color:#0563C1;">🔗 İlanı Gör</a>
            &nbsp;|&nbsp;<small>Ülke: {job['country']}</small>
            {f"<br><small>Yan Haklar: {job['benefit']}</small>" if job.get('benefit') else ""}
        </div>"""
    html_content = f"""
    <html><body style="font-family:Arial,sans-serif;max-width:700px;margin:auto;">
    <h2 style="color:#1B3A6B;">LinkedIn İş İlanları — {len(jobs)} ilan bulundu</h2>
    <p>Tarih: {datetime.now().strftime('%d.%m.%Y')} | Anahtar Kelime: <strong>{JOB_KEYWORD}</strong></p>
    {jobs_html}
    <hr><p style="color:#888;font-size:12px;">Otomatik Scraper • {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
    </body></html>"""
    try:
        msg            = MIMEMultipart()
        msg['From']    = sender
        msg['To']      = receiver
        msg['Subject'] = f"{len(jobs)} Yeni LinkedIn İlanı — {datetime.now().strftime('%d.%m.%Y')}"
        msg.attach(MIMEText(html_content, 'html'))
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender, password)
            server.send_message(msg)
        print(f"✅ E-posta {receiver} adresine gönderildi")
        return True
    except Exception as e:
        print("⚠️ E-posta gönderilemedi:", e)
        return False

# --- MAIN SCRAPING LOOP ---
all_jobs = []

for country in COUNTRIES:
    print(f"\n=== {country} için LinkedIn İlanları Taranıyor ===")
    url = build_linkedin_url(JOB_KEYWORD, country, EXPERIENCE_LEVELS, WORKPLACE_TYPES, DATE_POSTED)
    print(f"🔗 URL: {url}")
    driver.get(url)
    scroll_page(driver)

    soup      = BeautifulSoup(driver.page_source, "html.parser")
    job_cards = soup.find_all("div", class_="base-card")
    print(f"📋 {len(job_cards)} ilan kartı bulundu")

    for idx, card in enumerate(job_cards):
        a_tag        = card.find("a",    class_="base-card__full-link")
        job_url      = a_tag["href"].strip() if a_tag else ""
        sr_span      = a_tag.find("span", class_="sr-only") if a_tag else None
        job_title    = sr_span.text.strip() if sr_span else ""
        company_tag  = card.find("h4",  class_="base-search-card__subtitle")
        company_a    = company_tag.find("a") if company_tag else None
        company_name = company_a.text.strip()     if company_a else ""
        company_url  = company_a["href"].strip()  if company_a else ""
        loc_tag      = card.find("span", class_="job-search-card__location")
        location     = loc_tag.text.strip()  if loc_tag  else ""
        ben_tag      = card.find("span", class_="job-posting-benefits__text")
        benefit      = ben_tag.text.strip()  if ben_tag  else ""
        post_tag     = card.find("time", class_="job-search-card__listdate")
        posted       = post_tag.text.strip() if post_tag else ""

        print(f"🔍 ({idx+1}/{len(job_cards)}) {job_title}")
        job_description, company_description = fetch_job_details(job_url)

        all_jobs.append({
            "country":             country,
            "job_title":           job_title,
            "company_name":        company_name,
            "company_url":         company_url,
            "location":            location,
            "benefit":             benefit,
            "posted":              posted,
            "job_url":             job_url,
            "job_description":     job_description,
            "company_description": company_description,
        })

# --- SAVE TO EXCEL ---
if all_jobs:
    excel_file = f"linkedin_jobs_{safe_keyword}_{'_'.join([c.replace(' ','') for c in COUNTRIES])}_{safe_exp}_{safe_workplace}_{safe_date}.xlsx"
    save_to_excel(all_jobs, excel_file)
else:
    print("⚠️ Hiç ilan bulunamadı.")

# --- SEND EMAIL ---
send_job_email(all_jobs, SENDER_EMAIL, RECEIVER_EMAIL, EMAIL_PASSWORD)

driver.quit()
print("\n✅ Tamamlandı.")